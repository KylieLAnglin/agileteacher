# %%
import os

import pandas as pd
from openpyxl import load_workbook
import scipy

from agileteacher.library import start
from agileteacher.library import process_text

# %%
TEXT_FILE_PATH = start.CLEAN_DATA_PATH + "text.csv"
ID = "id_attempt"
COLUMNS = ["text_clean"]

# %%

df = pd.read_csv(TEXT_FILE_PATH).set_index(ID)
df = df[COLUMNS]

# %%

# %% Version 1
matrix = process_text.vectorize_text(
    df,
    text_col="text_clean",
    remove_stopwords=False,
    tfidf=False,
    lemma=False,
    lsa=False,
)

file = start.RESULTS_PATH + "Pilot Study/Cosine Replicability 1.xlsx"
wb = load_workbook(file)
ws = wb.active

col = 2
for main in list(matrix.index):
    row = 2
    for comp in list(matrix.index):
        dist = 1 - scipy.spatial.distance.cosine(matrix.loc[main], matrix.loc[comp])
        ws.cell(row=row, column=col).value = round(dist, 2)
        row = row + 1
    col = col + 1

wb.save(file)


process_text.what_words_matter(matrix, "jonathan1", "jonathan3", 10)
process_text.top_terms(matrix, "jonathan1", 10)
process_text.top_terms(matrix, "jonathan3", 10)

process_text.what_words_matter(matrix, "miles1", "miles3", 5)


# %% Version 2
matrix = process_text.vectorize_text(
    df,
    text_col="text_clean",
    remove_stopwords=True,
    tfidf=False,
    lemma=True,
    lsa=False,
)

file = start.RESULTS_PATH + "Pilot Study/Cosine Replicability 2.xlsx"
wb = load_workbook(file)
ws = wb.active

col = 2
for main in list(df.index):
    row = 2
    for comp in list(df.index):
        dist = 1 - scipy.spatial.distance.cosine(matrix.loc[main], matrix.loc[comp])
        ws.cell(row=row, column=col).value = round(dist, 2)
        row = row + 1
    col = col + 1

wb.save(file)

process_text.top_terms(matrix, "jonathan1", 10)
process_text.top_terms(matrix, "jonathan3", 10)


# %% Version 3
df["nostop"] = [
    process_text.process_text(
        text, lower_case=False, remove_punct=False, remove_stopwords=True
    )
    for text in df.text_clean
]


# matrix = process_text.doc_matrix_with_embeddings(df, "nostop").set_index("new_index")
tfidf = process_text.vectorize_text(
    df=df,
    text_col="text_clean",
    remove_stopwords=True,
    tfidf=True,
    lemma=True,
    lsa=False,
)

matrix = process_text.weighted_doc_matrix_with_embeddings(tfidf_matrix=tfidf).set_index(
    "id_attempt"
)

file = start.RESULTS_PATH + "Pilot Study/Cosine Replicability 3.xlsx"
wb = load_workbook(file)
ws = wb.active

col = 2
for main in list(df.index):
    row = 2
    for comp in list(df.index):
        dist = 1 - scipy.spatial.distance.cosine(matrix.loc[main], matrix.loc[comp])
        ws.cell(row=row, column=col).value = round(dist, 2)
        row = row + 1
    col = col + 1

wb.save(file)

# %%
