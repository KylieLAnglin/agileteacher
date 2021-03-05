# %%
import pandas as pd
from openpyxl import load_workbook


from agileteacher.library import start

# %%
OUTCOME_FILE_PATH = start.RAW_DATA_PATH + "pilot_study_outcomes.xlsx"
# ID = "id_attempt"
# COLUMNS = ["text_clean"]

# %%
df = pd.read_excel(OUTCOME_FILE_PATH)
df = df.reset_index().rename(columns={"index": "component"})
df = df[
    [
        "component",
        "esteban1",
        "esteban2",
        "esteban3",
        "jonathan1",
        "jonathan2",
        "jonathan3",
        "kaycie1",
        "kaycie2",
        "kaycie3",
        "samantha1",
        "samantha2",
        "samantha3",
        "kylie1",
        "kylie2",
        "kylie3",
        "miles1",
        "miles2",
        "miles3",
        "naomi1",
        "naomi2",
        "naomi3",
    ]
]
df["esteban1to2"] = abs(df.esteban1 - df.esteban2)
df["esteban2to3"] = abs(df.esteban2 - df.esteban3)
df["esteban1to3"] = abs(df.esteban1 - df.esteban3)

df["jonathan1to2"] = abs(df.jonathan1 - df.jonathan2)
df["jonathan2to3"] = abs(df.jonathan2 - df.jonathan3)
df["jonathan1to3"] = abs(df.jonathan1 - df.jonathan3)

df["kaycie1to2"] = abs(df.kaycie1 - df.kaycie2)
df["kaycie2to3"] = abs(df.kaycie2 - df.kaycie3)
df["kaycie1to3"] = abs(df.kaycie1 - df.kaycie3)

df["samantha1to2"] = abs(df.samantha1 - df.samantha2)
df["samantha2to3"] = abs(df.samantha2 - df.samantha3)
df["samantha1to3"] = abs(df.samantha1 - df.samantha3)

df["kylie1to2"] = abs(df.kylie1 - df.kylie2)
df["kylie2to3"] = abs(df.kylie2 - df.kylie3)
df["kylie1to3"] = abs(df.kylie1 - df.kylie3)

df["miles1to2"] = abs(df.miles1 - df.miles2)
df["miles2to3"] = abs(df.miles2 - df.miles3)
df["miles1to3"] = abs(df.miles1 - df.miles3)

df["naomi1to2"] = abs(df.naomi1 - df.naomi2)
df["naomi2to3"] = abs(df.naomi2 - df.naomi3)
df["naomi1to3"] = abs(df.naomi1 - df.naomi3)

df_sum = df.agg(["sum"]).filter(like="to")
df_sum = df_sum.T

# %%
df_sum.to_csv(start.CLEAN_DATA_PATH + "outcomes.csv")
# %%

file = start.RESULTS_PATH + "Pilot Study/Comparing Results.xlsx"
wb = load_workbook(file)
ws = wb.active


row = 3
for result in [
    "esteban1to2",
    "jonathan1to2",
    "kaycie1to2",
    "samantha1to2",
    "kylie1to2",
    "miles1to2",
    "naomi1to2",
]:
    col = 2
    ws.cell(row=row, column=col, value=int(df_sum.loc[result]))
    row = row + 1


row = 3
for result in [
    "esteban2to3",
    "jonathan2to3",
    "kaycie2to3",
    "samantha2to3",
    "kylie2to3",
    "miles2to3",
    "naomi2to3",
]:
    col = 4
    ws.cell(row=row, column=col, value=int(df_sum.loc[result]))
    row = row + 1


wb.save(file)


# %%

# %%
import os

import pandas as pd
from openpyxl import load_workbook
import scipy

from agileteacher.library import start
from agileteacher.library import process_text

# %%
TEXT_FILE_PATH = start.CLEAN_DATA_PATH + "text.csv"
ID = "id_attempt"
COLUMNS = ["text_clean"]

# %%

text_df = pd.read_csv(TEXT_FILE_PATH).set_index(ID)
text_df = text_df[COLUMNS]

matrix = process_text.vectorize_text(
    text_df,
    text_col="text_clean",
    remove_stopwords=True,
    tfidf=False,
    lemma=True,
    lsa=False,
)

# text_df["nostop"] = [
#     process_text.process_text(
#         text, lower_case=False, remove_punct=False, remove_stopwords=True
#     )
#     for text in text_df.text_clean
# ]


# # matrix = process_text.doc_matrix_with_embeddings(df, "nostop").set_index("new_index")
# tfidf = process_text.vectorize_text(
#     df=text_df,
#     text_col="text_clean",
#     remove_stopwords=True,
#     tfidf=True,
#     lemma=True,
#     lsa=False,
# )

# matrix = process_text.weighted_doc_matrix_with_embeddings(tfidf_matrix=tfidf).set_index(
#     "id_attempt"
# )


file = start.RESULTS_PATH + "Pilot Study/Comparing Results.xlsx"
wb = load_workbook(file)
ws = wb.active

ws.cell(row=3, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["esteban1"], matrix.loc["esteban2"]),
    2,
)
ws.cell(row=3, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["esteban2"], matrix.loc["esteban3"]),
    2,
)

ws.cell(row=4, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["jonathan1"], matrix.loc["jonathan2"]),
    2,
)
ws.cell(row=4, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["jonathan2"], matrix.loc["jonathan3"]),
    2,
)

ws.cell(row=5, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["kaycie1"], matrix.loc["kaycie2"]),
    2,
)
ws.cell(row=5, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["kaycie2"], matrix.loc["kaycie3"]),
    2,
)

ws.cell(row=6, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["samantha1"], matrix.loc["samantha2"]),
    2,
)
ws.cell(row=6, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["samantha2"], matrix.loc["samantha3"]),
    2,
)

ws.cell(row=7, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["kylie1"], matrix.loc["kylie2"]), 2
)
ws.cell(row=7, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["kylie2"], matrix.loc["kylie3"]), 2
)

ws.cell(row=8, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["miles1"], matrix.loc["miles2"]), 2
)
ws.cell(row=8, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["miles2"], matrix.loc["miles3"]), 2
)

ws.cell(row=9, column=3).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["naomi1"], matrix.loc["naomi2"]), 2
)
ws.cell(row=9, column=5).value = round(
    scipy.spatial.distance.euclidean(matrix.loc["naomi2"], matrix.loc["naomi3"]), 2
)


wb.save(file)
# %%
process_text.what_words_matter(matrix, "samantha1", "samantha2", 20)

# %%
